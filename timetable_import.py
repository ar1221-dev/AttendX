"""Robust timetable file parsing and import utilities."""

import csv
import re
from datetime import datetime, time as time_type
from io import BytesIO, StringIO

import openpyxl

DAY_MAP = {
    'monday': 0, 'mon': 0, 'mo': 0,
    'tuesday': 1, 'tue': 1, 'tu': 1,
    'wednesday': 2, 'wed': 2, 'we': 2,
    'thursday': 3, 'thu': 3, 'th': 3,
    'friday': 4, 'fri': 4, 'fr': 4,
    'saturday': 5, 'sat': 5, 'sa': 5,
    'sunday': 6, 'sun': 6, 'su': 6,
}

DAY_NAMES_SHORT = ('mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun')

SUBJECT_KEYS = (
    'subject', 'subject name', 'subject code', 'code', 'course', 'course name',
    'paper', 'unit', 'lecture', 'class', 'class name', 'module', 'topic',
    'paper name', 'subjectname', 'subjects',
)
DAY_KEYS = ('day', 'day of week', 'weekday', 'dow', 'days')
START_KEYS = ('start time', 'start_time', 'start', 'from', 'time from', 'begin', 'in time', 'from time')
END_KEYS = ('end time', 'end_time', 'end', 'to', 'time to', 'finish', 'out time', 'till', 'until')
TIME_RANGE_KEYS = ('time', 'timing', 'timings', 'slot', 'period', 'hours', 'class time', 'lecture time')
ROOM_KEYS = ('room', 'location', 'venue', 'hall', 'classroom', 'lab', 'block')
FACULTY_KEYS = ('faculty', 'teacher', 'professor', 'instructor', 'staff', 'tutor')
NOTES_KEYS = ('notes', 'note', 'remarks', 'comment')

HEADER_KEYWORDS = (
    'day', 'subject', 'time', 'course', 'start', 'end', 'room', 'faculty',
    'lecture', 'class', 'period', 'timing', 'slot', 'paper', 'module', 'hour',
)


def _cell_to_str(cell):
    if cell is None:
        return ''
    if isinstance(cell, datetime):
        return cell.strftime('%H:%M')
    if isinstance(cell, time_type):
        return cell.strftime('%H:%M')
    if isinstance(cell, float):
        if 0 <= cell < 1:
            total_minutes = int(round(cell * 24 * 60))
            h, m = divmod(total_minutes, 60)
            return f'{h:02d}:{m:02d}'
    return str(cell).strip()


def _normalize_header(h):
    h = str(h).strip().lower()
    h = re.sub(r'[^\w\s]', ' ', h)
    return re.sub(r'\s+', ' ', h).strip()


def _header_score(cells):
    """Score how likely a row is a header row (higher = more likely)."""
    normalized = [_normalize_header(c) for c in cells if c and str(c).strip()]
    if len(normalized) < 2:
        return 0
    score = 0
    for c in normalized:
        if c in DAY_KEYS or c in ('weekday', 'days'):
            score += 3
        elif any(c == k or c.startswith(k + ' ') or c.endswith(' ' + k) for k in SUBJECT_KEYS):
            score += 2
        elif any(c == k or 'start' in c or 'end' in c for k in START_KEYS + END_KEYS + TIME_RANGE_KEYS):
            score += 2
        elif c in ROOM_KEYS or c in FACULTY_KEYS:
            score += 1
    return score


def _find_header_index(all_rows, max_scan=25):
    """Find the best header row index within the first max_scan rows."""
    best_idx = None
    best_score = 0
    limit = min(len(all_rows), max_scan)
    for i in range(limit):
        if _row_looks_like_data(all_rows[i]):
            continue
        score = _header_score(all_rows[i])
        if score > best_score:
            best_score = score
            best_idx = i
    if best_score >= 2:
        return best_idx
    for i in range(limit):
        cells = all_rows[i]
        non_empty = [c for c in cells if c and str(c).strip()]
        if len(non_empty) >= 3 and not _row_looks_like_data(cells):
            return i
    return None


def _get_field(row, keys):
    for key in keys:
        if key in row and str(row[key]).strip():
            return str(row[key]).strip()
    for col, val in row.items():
        if not val or not col or col.startswith('_'):
            continue
        col_l = col.lower()
        for key in keys:
            if key in col_l or col_l in key:
                return str(val).strip()
    return ''


def parse_time_value(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.strftime('%H:%M')
    if isinstance(val, time_type):
        return val.strftime('%H:%M')
    if isinstance(val, float) and 0 <= val < 1:
        total_minutes = int(round(val * 24 * 60))
        h, m = divmod(total_minutes, 60)
        return f'{h:02d}:{m:02d}'

    s = str(val).strip()
    if ' ' in s and ':' in s:
        for part in s.split():
            if ':' in part:
                s = part
                break

    formats = (
        '%H:%M', '%H:%M:%S', '%I:%M %p', '%I:%M%p',
        '%I:%M:%S %p', '%I:%M:%S%p', '%H.%M', '%I:%M %P',
    )
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime('%H:%M')
        except ValueError:
            continue

    m = re.match(r'^(\d{1,2})[:\.](\d{2})\s*(am|pm)?$', s, re.I)
    if m:
        h, mi, ap = int(m.group(1)), int(m.group(2)), (m.group(3) or '').lower()
        if ap == 'pm' and h < 12:
            h += 12
        if ap == 'am' and h == 12:
            h = 0
        return f'{h:02d}:{mi:02d}'
    return None


def parse_time_range(val):
    """Parse '9:00-10:00', '9:00 to 10:00', '9:00 – 10:00'."""
    if not val:
        return None, None
    s = str(val).strip()
    parts = re.split(r'\s*(?:-|–|—|to)\s*', s, maxsplit=1, flags=re.I)
    if len(parts) == 2:
        start = parse_time_value(parts[0].strip())
        end = parse_time_value(parts[1].strip())
        if start and end:
            return start, end
    return None, None


def parse_day_value(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        n = int(val)
        if 0 <= n <= 6:
            return n
        if 1 <= n <= 7:
            return n - 1
    s = str(val).strip()
    if s.isdigit():
        n = int(s)
        if 0 <= n <= 6:
            return n
        if 1 <= n <= 7:
            return n - 1
    s_clean = re.sub(r'[^a-z0-9]', '', s.lower())
    if s_clean in DAY_MAP:
        return DAY_MAP[s_clean]
    for name, idx in DAY_MAP.items():
        if len(name) >= 3 and (s_clean.startswith(name[:3]) or name.startswith(s_clean[:3])):
            return idx
    return None


def _rows_to_dicts(header_cells, data_rows):
    header = [_normalize_header(c) or f'col_{i}' for i, c in enumerate(header_cells)]
    rows = []
    for raw in data_rows:
        if not any(str(c).strip() for c in raw):
            continue
        cells = [_cell_to_str(c) for c in raw]
        row = {header[i]: cells[i] if i < len(cells) else '' for i in range(len(header))}
        rows.append(row)
    return rows


def extract_time_range_str(s):
    """Extract a time range substring like 9-10 or 09:00-10:00 from a string."""
    pattern = r'\d{1,2}(?::\d{2})?\s*(?:am|pm)?\s*(?:-|–|—|to)\s*\d{1,2}(?::\d{2})?\s*(?:am|pm)?'
    m = re.search(pattern, s, re.I)
    if m:
        return m.group(0)
    return None


def parse_period_header(val):
    """Parse period labels like 9-10, 12-1, 1-2, or nested labels like '1 (9-10)' into start/end times."""
    if val is None or val == '':
        return None, None
    s = str(val).strip()
    tr_str = extract_time_range_str(s)
    if not tr_str:
        return None, None
    
    ts, te = parse_time_range(tr_str)
    if ts and te:
        return ts, te
    m = re.match(r'^(\d{1,2})\s*[-–—]\s*(\d{1,2})$', tr_str)
    if not m:
        return None, None
    sh, eh = int(m.group(1)), int(m.group(2))
    if eh < sh:  # 12-1 → 12:00 to 13:00
        return f'{sh:02d}:00', f'{eh + 12:02d}:00'
    if 1 <= sh <= 7 and 1 <= eh <= 8 and sh < eh:  # 1-2, 2-3, 3-4 PM
        return f'{sh + 12:02d}:00', f'{eh + 12:02d}:00'
    return f'{sh:02d}:00', f'{eh:02d}:00'


def _is_period_header(val):
    s = _cell_to_str(val)
    if not s:
        return False
    tr_str = extract_time_range_str(s)
    return tr_str is not None



def _try_day_row_grid_format(all_rows):
    """
    Parse grid timetables where DAYS are rows and TIME PERIODS are columns:
    |     | 9-10 | 10-11 | 11-12 | ...
    | Mo  | PA   | PA    | BEE   | ...
  """
    header_row_idx = None
    period_columns = {}

    for i, row in enumerate(all_rows[:20]):
        periods = {}
        for j, cell in enumerate(row):
            if j == 0:
                continue
            val = _cell_to_str(cell)
            if _is_period_header(val):
                t_start, t_end = parse_period_header(val)
                if t_start and t_end:
                    periods[j] = (t_start, t_end)
        if len(periods) >= 2:
            header_row_idx = i
            period_columns = periods
            break

    if header_row_idx is None:
        return []

    rows = []
    skip_subjects = {'', '-', 'na', 'n/a', 'free', 'break', 'lunch', 'none', 'x'}

    for row in all_rows[header_row_idx + 1:]:
        if not row:
            continue
        day_val = _cell_to_str(row[0])
        day_idx = parse_day_value(day_val)
        if day_idx is None:
            continue

        for col_idx, (t_start, t_end) in period_columns.items():
            if col_idx >= len(row):
                continue
            subject = _cell_to_str(row[col_idx])
            if not subject or subject.lower() in skip_subjects:
                continue
            rows.append({
                'day': str(day_idx),
                'subject': subject,
                'start time': t_start,
                'end time': t_end,
            })

    return rows


def _detect_matrix_header(row):
    """Return {col_index: day_of_week} if row has 2+ day column headers."""
    day_columns = {}
    for j, cell in enumerate(row):
        if not cell:
            continue
        day_idx = parse_day_value(cell)
        if day_idx is not None and j > 0:
            day_columns[j] = day_idx
    return day_columns if len(day_columns) >= 2 else None


def _row_looks_like_data(cells):
    """True if row appears to be a data row (day name in first column)."""
    if not cells:
        return False
    first = _cell_to_str(cells[0])
    return parse_day_value(first) is not None


def _parse_matrix_from_row(all_rows, header_idx, day_columns):
    rows = []
    for row in all_rows[header_idx + 1:]:
        if not any(c and str(c).strip() for c in row):
            continue
        time_val = _cell_to_str(row[0]) if row else ''
        t_start, t_end = parse_time_range(time_val)
        if not t_start:
            t_start = parse_time_value(time_val)
        if not t_start:
            continue
        if not t_end:
            try:
                h, m = map(int, t_start.split(':'))
                t_end = f'{(h + 1):02d}:{m:02d}'
            except ValueError:
                continue

        for col_idx, day_idx in day_columns.items():
            if col_idx >= len(row):
                continue
            subject = _cell_to_str(row[col_idx])
            if not subject or subject.lower() in ('-', 'na', 'n/a', 'free', 'break', 'lunch', 'none', ''):
                continue
            rows.append({
                'day': str(day_idx),
                'subject': subject,
                'start time': t_start,
                'end time': t_end,
            })
    return rows


def _try_matrix_format(all_rows):
    for i, row in enumerate(all_rows[:25]):
        day_columns = _detect_matrix_header(row)
        if day_columns:
            rows = _parse_matrix_from_row(all_rows, i, day_columns)
            if rows:
                return rows
    return []


def _try_positional_format(all_rows):
    """Fallback: col0=day, col1=subject, col2=start, col3=end (skip title rows)."""
    rows = []
    for raw in all_rows:
        cells = [_cell_to_str(c) for c in raw]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 3:
            continue
        day_idx = parse_day_value(cells[0])
        if day_idx is None:
            continue
        subject = cells[1] if len(cells) > 1 else ''
        if not subject:
            continue
        t_start = parse_time_value(cells[2]) if len(cells) > 2 else None
        t_end = parse_time_value(cells[3]) if len(cells) > 3 else None
        if not t_start:
            t_start, t_end = parse_time_range(cells[2])
        if not t_start or not t_end:
            continue
        rows.append({
            'day': str(day_idx),
            'subject': subject,
            'start time': t_start,
            'end time': t_end,
            'room': cells[4] if len(cells) > 4 else '',
        })
    return rows


def _extract_rows_from_sheet(all_rows):
    """Try multiple strategies to extract timetable rows."""
    if not all_rows:
        return [], 'empty'

    # Strategy 1: Day-rows grid (Mo/Tu/We rows × 9-10/10-11 columns) — most common college format
    day_grid_rows = _try_day_row_grid_format(all_rows)
    if day_grid_rows:
        return day_grid_rows, 'day_grid'

    # Strategy 2: Matrix/grid format (days as columns, time in rows)
    matrix_rows = _try_matrix_format(all_rows)
    if matrix_rows:
        return matrix_rows, 'matrix'

    # Strategy 3: Standard header + data rows
    header_idx = _find_header_index(all_rows)
    if header_idx is not None:
        header_cells = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1:]
        rows = _rows_to_dicts(header_cells, data_rows)
        if rows:
            return rows, 'header'

    # Strategy 4: Positional columns without clear header
    positional = _try_positional_format(all_rows)
    if positional:
        return positional, 'positional'

    return [], 'none'


def read_csv_rows(file_bytes):
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(StringIO(text, newline=''))
    all_rows = []
    for raw in reader:
        if any(str(c).strip() for c in raw):
            all_rows.append(raw)
    rows, _ = _extract_rows_from_sheet(all_rows)
    return rows


def read_xlsx_rows(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    all_rows = []
    for sheet in wb.worksheets:
        sheet_rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                sheet_rows.append(list(row))
        rows, method = _extract_rows_from_sheet(sheet_rows)
        if rows:
            return rows
        if not all_rows and sheet_rows:
            all_rows = sheet_rows
    # Last attempt on first sheet raw data
    rows, _ = _extract_rows_from_sheet(all_rows)
    return rows


    return rows


def parse_leaves_file(file_obj, filename):
    """Parse a CSV/XLSX file of leave dates."""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext == 'csv':
        text = file_obj.read().decode('utf-8-sig', errors='replace')
        reader = csv.reader(StringIO(text, newline=''))
        all_rows = [r for r in reader if any(str(c).strip() for c in r)]
    elif ext in ('xlsx', 'xlsm'):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        all_rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                all_rows.append([_cell_to_str(c) for c in row])
    else:
        raise ValueError('Use .csv or .xlsx file')

    if not all_rows:
        return []

    # Find header row or use first row as header
    header_idx = 0
    for i, row in enumerate(all_rows[:5]):
        joined = ' '.join(str(c).lower() for c in row)
        if 'date' in joined or 'leave' in joined or 'holiday' in joined:
            header_idx = i
            break

    header = [_normalize_header(c) or f'col_{j}' for j, c in enumerate(all_rows[header_idx])]
    rows = []
    for raw in all_rows[header_idx + 1:]:
        if not any(str(c).strip() for c in raw):
            continue
        row = {}
        for i, key in enumerate(header):
            row[key] = str(raw[i]).strip() if i < len(raw) else ''
        rows.append(row)

    # If no data rows, try single-column dates without header
    if not rows and len(all_rows) >= 1:
        for raw in all_rows:
            for cell in raw:
                cell = str(cell).strip()
                if cell and (re.match(r'\d{4}-\d{2}-\d{2}', cell) or re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{4}', cell)):
                    rows.append({'date': cell, 'reason': 'Leave'})
    return rows


def parse_timetable_file(file_obj, filename):
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext == 'csv':
        data = file_obj.read()
        rows = read_csv_rows(data)
    elif ext in ('xlsx', 'xlsm'):
        rows = read_xlsx_rows(file_obj)
    elif ext == 'xls':
        raise ValueError('Old .xls format not supported. Please save as .xlsx in Excel and re-upload.')
    else:
        raise ValueError('Unsupported file format. Use .csv or .xlsx')
    return rows


def import_timetable_rows(db, semester_id, version_id, rows):
    subjects_by_code = {s['code'].lower(): s['id'] for s in db.get_subjects(semester_id) if s['code']}
    subjects_by_name = {s['name'].lower(): s['id'] for s in db.get_subjects(semester_id)}

    success = 0
    skipped = {'incomplete': 0, 'invalid_day': 0, 'invalid_time': 0}

    for r in rows:
        day_raw = _get_field(r, DAY_KEYS)
        if not day_raw and 'day' in r:
            day_raw = str(r['day'])
        subject_str = _get_field(r, SUBJECT_KEYS)
        start_raw = _get_field(r, START_KEYS)
        end_raw = _get_field(r, END_KEYS)
        room = _get_field(r, ROOM_KEYS)
        faculty = _get_field(r, FACULTY_KEYS)
        notes = _get_field(r, NOTES_KEYS)

        # Combined time column e.g. "9:00-10:00"
        if not start_raw or not end_raw:
            time_range = _get_field(r, TIME_RANGE_KEYS)
            if time_range:
                ts, te = parse_time_range(time_range)
                if ts:
                    start_raw = ts
                if te:
                    end_raw = te

        if not day_raw or not subject_str:
            skipped['incomplete'] += 1
            continue

        day_idx = parse_day_value(day_raw)
        if day_idx is None:
            skipped['invalid_day'] += 1
            continue

        t_start = parse_time_value(start_raw)
        t_end = parse_time_value(end_raw)
        if not t_start or not t_end:
            skipped['invalid_time'] += 1
            continue

        sub_key = subject_str.lower()
        if sub_key == 'bee eval':
            sub_key = 'bee'
            
        if sub_key in subjects_by_code:
            subject_id = subjects_by_code[sub_key]
        elif sub_key in subjects_by_name:
            subject_id = subjects_by_name[sub_key]
        else:
            # If the subject is normalized to 'bee', use 'BEE' as the name, otherwise subject_str
            final_name = 'BEE' if sub_key == 'bee' else subject_str
            code = ''.join(w[0] for w in final_name.split() if w).upper()[:8] or final_name[:8].upper()
            subject_id = db.add_subject(semester_id, final_name, code, faculty or '', 4)
            subjects_by_code[sub_key] = subject_id
            subjects_by_name[sub_key] = subject_id

        db.add_timetable_entry(version_id, subject_id, day_idx, t_start, t_end, room, notes)
        success += 1

    return {'success': success, 'skipped': skipped, 'total_rows': len(rows)}
